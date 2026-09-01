"""Import transactions, categories and assets from a CSV or Excel file.

Reusable service: reads a file (CSV or Excel), normalizes its rows, and
creates categories + a default Liquid asset + transactions, all scoped to
the given user. Used by the /api/import endpoint.

Parsing is separated from persistence: `_parse_csv`/`_parse_excel` return
normalized rows with no DB access, so a preview can reuse the exact same
parsing without writing anything.
"""

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from pynance.models.asset import Asset
from pynance.models.category import Category
from pynance.models.transaction import Transaction
from pynance.models.types import AssetType, TransactionType

TRANSACTIONS_SHEET = "List_of_incomes_expenditures"
CATEGORIES_SHEET = "conditional_setting (dont see i"
DEFAULT_ASSET_NAME = "Liquid"


@dataclass(frozen=True)
class ImportRow:
    description: str
    occurred_on: date
    transaction_type: TransactionType
    category: str
    amount: Decimal


@dataclass
class ImportResult:
    categories_created: int
    transactions_imported: int
    skipped: int


def parse_amount(value: object) -> Decimal:
    """Accept int/float, dot-decimal ('2.50'), or European-format ('1.234,56', '48,6')."""
    if isinstance(value, str):
        s = value.strip()
        if "," in s:
            # European: 1.234,56 -> 1234.56 ; 48,6 -> 48.6
            s = s.replace(".", "").replace(",", ".")
        return Decimal(s).quantize(Decimal("0.01"))
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _transaction_type(value: str) -> TransactionType:
    normalized = value.strip().lower()
    return TransactionType.INCOME if normalized == "income" else TransactionType.EXPENSE


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()


def _parse_csv(content: bytes) -> tuple[list[ImportRow], int]:
    """Parse CSV into rows. Expected header: Details,Date,Type,Category,Amount (EUR)."""
    import csv
    import io

    rows: list[ImportRow] = []
    skipped = 0
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))

    for record in reader:
        description = (record.get("Details") or "").strip()
        date_value = record.get("Date")
        type_value = record.get("Type")
        category = record.get("Category")
        amount = record.get("Amount (EUR)") or record.get("Amount")
        if not (description and date_value and type_value and category and amount):
            skipped += 1
            continue
        try:
            rows.append(
                ImportRow(
                    description=description,
                    occurred_on=_parse_date(date_value),
                    transaction_type=_transaction_type(type_value),
                    category=category.strip(),
                    amount=parse_amount(amount),
                )
            )
        except ValueError, InvalidOperation:
            skipped += 1
    return rows, skipped


def _category_pairs_from_workbook(wb: Any) -> list[tuple[str, str]]:
    """Collect (type, category) pairs from the conditional_setting sheet, if present."""
    category_pairs: list[tuple[str, str]] = []
    if CATEGORIES_SHEET not in wb.sheetnames:
        return category_pairs
    ws = wb[CATEGORIES_SHEET]
    for row in ws.iter_rows(values_only=True):
        if len(row) < 2 or not row[0] or not row[1]:
            continue
        sheet_type, sheet_category = row[0], row[1]
        if sheet_type == "Type" or sheet_category == "Category":
            continue
        sheet_type = str(sheet_type).strip()
        sheet_category = str(sheet_category).strip()
        if sheet_type not in ("Income", "Expenditure"):
            continue
        category_pairs.append((sheet_type, sheet_category))
    return category_pairs


def _parse_excel(content: bytes) -> tuple[list[ImportRow], int]:
    """Parse the transactions sheet into rows (no DB access)."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet_name = TRANSACTIONS_SHEET if TRANSACTIONS_SHEET in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows: list[ImportRow] = []
    skipped = 0
    for raw in ws.iter_rows(values_only=True):
        if len(raw) < 6:
            continue
        details: object | None = raw[1]
        date_value: object | None = raw[2]
        type_value: object | None = raw[3]
        category: object | None = raw[4]
        amount: object = raw[5]
        if (
            details is None
            or details == "Details"
            or type_value is None
            or category is None
            or date_value is None
        ):
            continue
        try:
            rows.append(
                ImportRow(
                    description=str(details).strip(),
                    occurred_on=_parse_date(date_value),
                    transaction_type=_transaction_type(str(type_value)),
                    category=str(category).strip(),
                    amount=parse_amount(amount),
                )
            )
        except ValueError, InvalidOperation:
            skipped += 1
    return rows, skipped


def _parse_file(filename: str, content: bytes) -> tuple[list[ImportRow], int]:
    """Dispatch to the right parser based on the file extension."""
    name = filename.lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith(".xlsx"):
        return _parse_excel(content)
    raise ValueError("Unsupported file type. Use CSV or Excel (.csv, .xlsx).")


def _get_or_create_category(
    db: Session, user_id: int, name: str, transaction_type: TransactionType
) -> Category:
    existing = db.execute(
        select(Category).where(Category.user_id == user_id, Category.name == name)
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    category = Category(name=name, transaction_type=transaction_type, user_id=user_id)
    db.add(category)
    db.flush()
    return category


def _get_or_create_liquid(db: Session, user_id: int) -> Asset:
    existing = db.execute(
        select(Asset).where(Asset.user_id == user_id, Asset.asset_type == AssetType.LIQUID)
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    asset = Asset(name=DEFAULT_ASSET_NAME, asset_type=AssetType.LIQUID, user_id=user_id)
    db.add(asset)
    db.flush()
    return asset


def _import_rows(db: Session, user_id: int, rows: list[ImportRow]) -> ImportResult:
    liquid = _get_or_create_liquid(db, user_id)
    categories_created = 0
    transactions_imported = 0
    skipped = 0

    for row in rows:
        existing = db.execute(
            select(Category).where(Category.user_id == user_id, Category.name == row.category)
        ).scalar_one_or_none()
        if existing is None:
            categories_created += 1
        category = _get_or_create_category(db, user_id, row.category, row.transaction_type)

        db.add(
            Transaction(
                amount=row.amount,
                category_id=category.id,
                asset_id=liquid.id,
                description=row.description,
                occurred_on=row.occurred_on,
                user_id=user_id,
            )
        )
        transactions_imported += 1

    db.commit()
    return ImportResult(categories_created, transactions_imported, skipped)


def import_csv(db: Session, user_id: int, content: bytes) -> ImportResult:
    rows, skipped = _parse_csv(content)
    result = _import_rows(db, user_id, rows)
    result.skipped = skipped
    return result


def import_excel(db: Session, user_id: int, content: bytes) -> ImportResult:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    category_pairs = _category_pairs_from_workbook(wb)
    for sheet_type, sheet_category in category_pairs:
        _get_or_create_category(db, user_id, sheet_category, _transaction_type(sheet_type))

    rows, skipped = _parse_excel(content)
    result = _import_rows(db, user_id, rows)
    result.skipped = skipped
    return result


def import_file(db: Session, user_id: int, filename: str, content: bytes) -> ImportResult:
    rows, skipped = _parse_file(filename, content)
    result = _import_rows(db, user_id, rows)
    result.skipped = skipped
    return result


def preview_file(filename: str, content: bytes, limit: int = 10) -> list[dict[str, object]]:
    """Parse a file and return the first `limit` rows, without writing to the DB."""
    rows, _ = _parse_file(filename, content)
    return [
        {
            "description": row.description,
            "occurred_on": row.occurred_on.isoformat(),
            "transaction_type": row.transaction_type.value,
            "category": row.category,
            "amount": str(row.amount),
        }
        for row in rows[:limit]
    ]
